import csv
import io
import json
from pathlib import Path
from typing import Any
from uuid import UUID

from backend.services.knowledge.types import KnowledgeObject


class KnowledgeIngestion:
    """Ingests documents from multiple sources into KnowledgeObject format."""

    SUPPORTED_EXTENSIONS = {
        ".pdf", ".docx", ".txt", ".csv", ".xlsx", ".xls",
        ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".json",
    }

    def from_text(
        self,
        title: str,
        text: str,
        document_type: str = "general",
        source: str = "upload",
        metadata: dict[str, Any] | None = None,
        structured_fields: dict[str, Any] | None = None,
        tables: list[dict] | None = None,
        tags: list[str] | None = None,
    ) -> KnowledgeObject:
        return KnowledgeObject(
            document_id=None,
            title=title,
            document_type=document_type,
            source=source,
            metadata=metadata or {},
            raw_text=text,
            structured_fields=structured_fields or {},
            tables=tables or [],
            tags=tags or [],
        )

    def from_ocr_result(self, ocr_knowledge: dict[str, Any]) -> KnowledgeObject:
        structured = ocr_knowledge.get("structured_content", {})
        fields = structured.get("fields", ocr_knowledge.get("fields", {}))
        doc_id_raw = ocr_knowledge.get("metadata", {}).get("document_id")
        document_id = None
        if doc_id_raw:
            try:
                document_id = UUID(str(doc_id_raw))
            except ValueError:
                pass

        return KnowledgeObject(
            document_id=document_id,
            title=ocr_knowledge.get("title", "OCR Document"),
            document_type=ocr_knowledge.get("document_class", structured.get("document_class", "general")),
            source=ocr_knowledge.get("metadata", {}).get("source", "ocr_pipeline"),
            metadata=ocr_knowledge.get("metadata", {}),
            raw_text=ocr_knowledge.get("raw_text", ""),
            structured_fields=fields,
            tables=ocr_knowledge.get("tables", structured.get("tables", [])),
            confidence=structured.get("confidence", {}).get("document"),
            tags=self._infer_tags(fields),
        )

    def from_file_bytes(
        self,
        file_name: str,
        data: bytes,
        source: str = "upload",
    ) -> KnowledgeObject:
        ext = Path(file_name).suffix.lower()
        title = Path(file_name).stem

        if ext == ".txt":
            return self.from_text(title, data.decode("utf-8", errors="replace"), source=source)
        if ext == ".csv":
            return self._from_csv(title, data, source)
        if ext == ".json":
            return self._from_json(title, data, source)
        if ext == ".docx":
            return self._from_docx(title, data, source)
        if ext in (".xlsx", ".xls"):
            return self._from_excel(title, data, source)

        return self.from_text(
            title,
            f"[Binary file: {file_name}] Use OCR pipeline for image/PDF processing.",
            document_type="general",
            source=source,
            metadata={"file_name": file_name, "requires_ocr": True},
        )

    def from_database_record(
        self,
        title: str,
        record: dict[str, Any],
        record_type: str,
        source: str = "database",
    ) -> KnowledgeObject:
        text = json.dumps(record, indent=2, default=str)
        return self.from_text(
            title=title,
            text=text,
            document_type=record_type,
            source=source,
            structured_fields=record,
            metadata={"ingestion": "database_record"},
        )

    def from_tally_export(self, export_data: dict[str, Any], source: str = "tally") -> KnowledgeObject:
        ledgers = export_data.get("ledgers", [])
        vouchers = export_data.get("vouchers", [])
        text_parts = [f"Tally Export\nLedgers: {len(ledgers)}\nVouchers: {len(vouchers)}"]
        for v in vouchers[:50]:
            text_parts.append(json.dumps(v, default=str))

        return self.from_text(
            title=export_data.get("company_name", "Tally Export"),
            text="\n".join(text_parts),
            document_type="ledger",
            source=source,
            structured_fields=export_data,
            tables=[{"table_type": "vouchers", "headers": list(vouchers[0].keys()) if vouchers else [], "rows": [[str(v.get(h, "")) for h in vouchers[0].keys()] for v in vouchers[:100]]}] if vouchers else [],
            metadata={"ingestion": "tally_export"},
            tags=["tally", "accounting"],
        )

    def _from_csv(self, title: str, data: bytes, source: str) -> KnowledgeObject:
        text = data.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        table = {"table_type": "csv", "headers": headers, "rows": data_rows}
        return self.from_text(title, text, document_type="general", source=source, tables=[table])

    def _from_json(self, title: str, data: bytes, source: str) -> KnowledgeObject:
        parsed = json.loads(data.decode("utf-8"))
        text = json.dumps(parsed, indent=2, default=str)
        fields = parsed if isinstance(parsed, dict) else {}
        return self.from_text(title, text, source=source, structured_fields=fields)

    def _from_docx(self, title: str, data: bytes, source: str) -> KnowledgeObject:
        from docx import Document as DocxDocument
        doc = DocxDocument(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)
        return self.from_text(title, text, source=source, metadata={"format": "docx"})

    def _from_excel(self, title: str, data: bytes, source: str) -> KnowledgeObject:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        tables = []
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [[str(cell.value or "") for cell in row] for row in ws.iter_rows(max_row=500)]
            if rows:
                tables.append({"table_type": "excel", "headers": rows[0], "rows": rows[1:], "sheet": sheet_name})
                text_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(" | ".join(r) for r in rows[:50]))
        return self.from_text(title, "\n\n".join(text_parts), source=source, tables=tables)

    def _infer_tags(self, fields: dict[str, Any]) -> list[str]:
        tags = []
        if fields.get("gstin"):
            tags.append("gst")
        if fields.get("invoice_number"):
            tags.append("invoice")
        if fields.get("vendor_name"):
            tags.append("vendor")
        if fields.get("customer_name"):
            tags.append("customer")
        return tags
